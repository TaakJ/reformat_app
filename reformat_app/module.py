import logging
from pathlib import Path
from os.path import join
import os
import glob
import shutil
import pandas as pd
import numpy as np
import openpyxl
import chardet
from io import StringIO
import xlrd
import csv
from .exception import CustomException
from .setup import Folder

class Convert2File:

    async def check_source_file(self) -> None:
        
        logging.info("Check Source file")

        _log = []
        for input_dir in self.input_dir:

            status = "not_found"
            if glob.glob(input_dir, recursive=True):
                status = "found"

            record = {"module": self.module,
                    "input_dir": input_dir,
                    "full_target": self.full_target,
                    "function": "check_source_file",
                    "status": status,
                    }
            _log.append(record)
            logging.info(f'Source file: "{input_dir}", Status: "{status}"')
            
        self.logging = _log
        
    async def retrieve_data_from_source_file(self) -> None:

        logging.info("Retrieve Data from Source file")

        for i, record in enumerate(self.logging):
            record.update({"function": "retrieve_data_from_source_file"})

            input_dir = record["input_dir"]
            types = Path(input_dir).suffix
            status_file = record["status"]
            try:
                if status_file == "found":
                    if [".xlsx", ".xls"].__contains__(types):
                        logging.info(f'Read Excel file: "{input_dir}"')
                        data = self.read_excel_file(i)
                    else:
                        logging.info(f'Read Text file: "{input_dir}"')
                        data = self.read_text_file(i)
                else:
                    raise FileNotFoundError(f'File Not Found: "{input_dir}"')

                status = "succeed"
                record.update({"data": data, "status": status})

            except Exception as err:
                record.update({"err": err})

            if "err" in record:
                raise CustomException(err=self.logging)

    def read_text_file(self, i: int) -> any:

        self.logging[i].update({"function": "read_text_file"})
        input_dir = self.logging[i]["input_dir"]
        try:
            file = open(input_dir, "rb")
            encoded = chardet.detect(file.read())["encoding"]
            file.seek(0)
            line = StringIO(file.read().decode(encoded))
            data = self.get_extract_data(i, line)

        except Exception as err:
            raise Exception(err)
        
        return data

    def read_excel_file(self, i: int) -> any:

        self.logging[i].update({"function": "read_excel_file"})
        input_dir = self.logging[i]["input_dir"]
        try:
            workbook = xlrd.open_workbook(input_dir)
            data = self.get_extract_data(i, workbook)

        except Exception as err:
            raise Exception(err)
        
        return data

    def initial_data_type(self, df: pd.DataFrame) -> pd.DataFrame:

        status = "failed"
        self.logging[-1].update({"function": "initial_data_type", "status": status})
        try:
            df = df.astype({
                    "ApplicationCode": object,
                    "AccountOwner": object,
                    "AccountName": object,
                    "AccountType": object,
                    "EntitlementName": object,
                    "SecondEntitlementName": object,
                    "ThirdEntitlementName": object,
                    "AccountStatus": object,
                    "IsPrivileged": object,
                    "AccountDescription": object,
                    "CreateDate": "datetime64[ms]",
                    "LastLogin": "datetime64[ms]",
                    "LastUpdatedDate": "datetime64[ms]",
                    "AdditionalAttribute": object})
            df[["CreateDate", "LastLogin", "LastUpdatedDate"]] = df[["CreateDate", "LastLogin", "LastUpdatedDate"]].apply(pd.to_datetime, format="%Y%m%d%H%M%S")

            if "remark" in df.columns:
                df = df.loc[df["remark"] != "Remove"]
            else:
                df["remark"] = "Insert"

        except Exception as err:
            raise Exception(err)

        status = "succeed"
        self.logging[-1].update({"status": status})
        
        return df

    def validate_data_change(self, df: pd.DataFrame, change_df: pd.DataFrame) -> dict:

        logging.info("Validate Data Change")
        self.change_rows = {}
        self.remove_rows = []
        
        status = "failed"
        self.logging[-1].update({"function": "validate_data_change", "status": status})

        ## set format record
        def format_record(record):
            return "\n".join("{!r} => {!r},".format(columns, values) for columns, values in record.items())

        if len(df.index) > len(change_df.index):
            self.remove_rows = [idx for idx in list(df.index) if idx not in list(change_df.index)]

        try:
            ## merge index.
            merge_index = np.union1d(df.index, change_df.index)

            ## as starter dataframe for compare
            df = df.reindex(index=merge_index, columns=df.columns).iloc[:, :-1]

            ## change data / new data
            change_df = change_df.reindex(index=merge_index, columns=change_df.columns).iloc[:, :-1]

            ## compare data
            df["count"] = pd.DataFrame(np.where(df.ne(change_df), True, df), index=df.index, columns=df.columns).apply(lambda x: (x == True).sum(), axis=1)

            i = 0
            for idx, row in enumerate(merge_index, 2):
                if row not in self.remove_rows:
                    record = {}
                    for data, change_data in zip(df.items(), change_df.items()):
                        ## No Change
                        if df.loc[row, "count"] != 15:
                            if df.loc[row, "count"] < 1:
                                df.loc[row, data[0]] = data[1][row]
                                df.loc[row, "remark"] = "No_change"
                            else:
                                ## Update
                                if data[1][row] != change_data[1][row]:
                                    record.update({data[0]: change_data[1][row]})
                                df.loc[row, data[0]] = change_data[1][row]
                                df.loc[row, "remark"] = "Update"
                        else:
                            ## Insert
                            record.update({data[0]: change_data[1][row]})
                            df.loc[row, data[0]] = change_data[1][row]
                            df.loc[row, "remark"] = "Insert"

                    if record != {}:
                        self.change_rows[idx] = format_record(record)
                else:
                    ## Remove
                    self.remove_rows[i] = idx
                    df.loc[row, "remark"] = "Remove"
                    i += 1

            ## set dataframe.
            df = df.drop(["count"], axis=1)
            rows = 2
            df.index += rows
            data_dict = df.to_dict(orient="index")

        except Exception as err:
            raise Exception(err)

        status = "succeed"
        self.logging[-1].update({"status": status})

        return data_dict

    async def write_data_to_tmp_file(self) -> None:

        logging.info("Write Data to Tmp file")

        status = "failed"
        for record in self.logging:
            try:
                if record["module"] == "Target_file":
                    try:
                        data = record["data"]
                        change_df = pd.DataFrame(data)
                        change_df = self.initial_data_type(change_df)

                        ## read tmp file
                        tmp_dir = join(Folder.TMP, self.module, self.date.strftime("%Y%m%d"))
                        os.makedirs(tmp_dir, exist_ok=True)
                        tmp_name = f"TMP_{Path(self.full_target).stem}.xlsx"
                        full_tmp = join(tmp_dir, tmp_name)
                        record.update({"input_dir": full_tmp})

                        ## crate tmp file
                        self.create_workbook()

                        ## set dataframe from tmp file
                        data = self.sheet.values
                        columns = next(data)[0:]
                        tmp_df = pd.DataFrame(data, columns=columns)
                        tmp_df = self.initial_data_type(tmp_df)

                        ## validate data change row by row
                        data_dict = self.validate_data_change(tmp_df, change_df)

                        ## write tmp file
                        status = self.write_worksheet(data_dict)

                    except Exception as err:
                        raise Exception(err)

                    record.update({"function": "write_data_to_tmp_file", "status": status})
                    logging.info(f'Write Data to Tmp file status: "{status}"')

            except Exception as err:
                record.update({"err": err})

            if "err" in record:
                raise CustomException(err=self.logging)

    def create_workbook(self) -> None:

        full_tmp = self.logging[-1]["input_dir"]
        logging.info(f'Create Tmp file: "{full_tmp}"')

        status = "failed"
        self.logging[-1].update({"function": "create_workbook", "status": status})

        try:
            self.create = False
            self.workbook = openpyxl.load_workbook(full_tmp)
            get_sheet = self.workbook.get_sheet_names()
            self.sheet_num = len(get_sheet)
            self.sheet_name = f"RUN_TIME_{self.sheet_num - 1}"

            if self.sheet_name in get_sheet:
                self.create = True
                self.sheet = self.workbook.get_sheet_by_name(self.sheet_name)
            else:
                self.sheet = self.workbook.get_sheet_by_name("Field Name")

        except FileNotFoundError:
            ## move from template file to tmp file
            template_name = "Application Data Requirements.xlsx"
            full_template = join(Folder.TEMPLATE, template_name)
            try:
                if not glob.glob(full_tmp, recursive=True):
                    shutil.copy2(full_template, full_tmp)
            except:
                raise

            self.workbook = openpyxl.load_workbook(full_tmp)
            self.sheet = self.workbook.get_sheet_by_name("Field Name")
            self.sheet_name = "RUN_TIME_1"
            self.sheet_num = 1

        status = "succeed"
        self.logging[-1].update({"status": status})

    def write_worksheet(self, change_data: dict) -> str:

        status = "failed"
        if self.create:
            self.sheet_name = f"RUN_TIME_{self.sheet_num}"
            self.sheet = self.workbook.create_sheet(self.sheet_name)

        logging.info(f"Write to {self.sheet}")
        
        rows = 2
        max_row = max(change_data, default=0)
        self.logging[-1].update({"function": "write_worksheet", "sheet_name": self.sheet_name, "status": status})
        try:
            # write column
            for idx, col in enumerate(change_data[rows].keys(), 1):
                self.sheet.cell(row=1, column=idx).value = col

            ## write row
            while rows <= max_row:
                for idx, col in enumerate(change_data[rows].keys(), 1):

                    if col in ["CreateDate", "LastLogin", "LastUpdatedDate"]:
                        change_data[rows][col] = change_data[rows][col].strftime("%Y%m%d%H%M%S")
                    self.sheet.cell(row=rows, column=idx).value = change_data[rows][col]

                    if col == "remark":
                        if rows in self.remove_rows:
                            ## Remove row
                            write_row = (f'{change_data[rows][col]} Rows: "{rows}" in Tmp file')
                        elif rows in self.change_rows.keys():
                            ## Update / Insert row
                            write_row = f'{change_data[rows][col]} Rows: "{rows}" in Tmp file\nUpdating records: {self.change_rows[rows]}'
                        else:
                            ## No change row
                            write_row = f'No Change Rows: "{rows}" in Tmp file'
                        logging.info(write_row)
                rows += 1

        except KeyError as err:
            raise KeyError(f"Can not Write rows: {err} in Tmp file")

        ## save file
        full_tmp = self.logging[-1]["input_dir"]
        self.sheet.title = self.sheet_name
        self.workbook.active = self.sheet
        self.workbook.move_sheet(self.workbook.active, offset=-self.sheet_num)
        self.workbook.save(full_tmp)

        status = "succeed"
        self.logging[-1].update({"status": status})
        
        return status

    async def write_data_to_target_file(self) -> None:

        logging.info("Write Data to Target file")
        
        status = "failed"
        for record in self.logging:
            try:

                if record["module"] == "Target_file":
                    try:
                        ## read tmp file or dataframe
                        if self.store_tmp is True:
                            full_tmp = record["input_dir"]
                            sheet_name = record["sheet_name"]
                            change_df = pd.read_excel(full_tmp, sheet_name=sheet_name, dtype=object)
                        else:
                            data = record["data"]
                            change_df = pd.DataFrame(data)
                        change_df = self.initial_data_type(change_df)
            
                        ## read csv file
                        target_df = self.read_csv()
                        
                        ## optimize data
                        data = self.optimize_data(target_df, change_df)
                        
                        ## write csv file
                        status = self.write_csv(self.full_target, data)

                    except Exception as err:
                        raise Exception(err)

                    record.update({"function": "write_data_to_target_file", "state": status})
                    logging.info(f'Write to Target file status: "{status}"')

            except Exception as err:
                record.update({"err": err})

        if "err" in record:
            raise CustomException(err=self.logging)

    def read_csv(self) -> pd.DataFrame:

        logging.info(f'Read Target file: "{self.full_target}"')

        status = "failed"
        self.logging[-1].update({"input_dir": self.full_target, "function": "read_csv", "status": status})

        try:
            data = []
            with open(self.full_target, "r", newline="\n") as reader:
                csv_reader = csv.reader(reader,
                                        skipinitialspace=True,
                                        delimiter=",",
                                        quotechar='"',
                                        quoting=csv.QUOTE_NONE)

                header = next(csv_reader)
                for row in csv_reader:
                    data.append(row)
            target_df = pd.DataFrame(data, columns=header)

        except FileNotFoundError:
            ## move from template file to target file
            template_name = join(Folder.TEMPLATE, "Application Data Requirements.xlsx")
            target_df = pd.read_excel(template_name)
            target_df.to_csv(self.full_target, index=None, header=True, sep=",")

        status = "succeed"
        self.logging[-1].update({"status": status})

        return target_df

    def optimize_data(self, target_df: pd.DataFrame, change_df: pd.DataFrame) -> dict:

        logging.info("Optimize Data Before Write to Target")

        status = "failed"
        self.logging[-1].update({"function": "optimize_data", "status": status})

        data = {}
        try:
            target_df = self.initial_data_type(target_df)

            ## filter data on batch date => DataFrame
            batch_df = target_df[target_df["CreateDate"].isin(np.array([pd.Timestamp(self.batch_date)]))].reset_index(drop=True)

            ## Validate data change row by row
            data_dict = self.validate_data_change(batch_df, change_df)

            ## filter data not on batch date => dict
            merge_data = (target_df[~target_df["CreateDate"].isin(np.array([pd.Timestamp(self.batch_date)]))].iloc[:, :-1].to_dict("index"))

            ## merge data from new and old data
            max_rows = max(merge_data, default=0)
            for idx, values in data_dict.items():
                if idx in self.change_rows or idx in self.remove_rows:
                    values.update({"mark_row": idx})
                merge_data = {**merge_data, **{max_rows + idx: values}}

            ## sorted order data on batch date
            i = 0
            for idx, values in enumerate(sorted(merge_data.values(), key=lambda d: d["CreateDate"]), 2):
                if "mark_row" in values.keys():
                    if values["mark_row"] in self.change_rows:
                        self.change_rows[idx] = self.change_rows.pop(values["mark_row"])
                    else:
                        self.remove_rows[i] = idx
                        i += 1
                    values.pop("mark_row")
                data.update({idx: values})

        except Exception as err:
            raise Exception(err)

        status = "succeed"
        self.logging[-1].update({"status": status})

        return data

    def write_csv(self, target_name: str, data: dict) -> str:

        logging.info(f'Write mode: "{self.write_mode}" in Target file: "{target_name}"')

        status = "failed"
        self.logging[-1].update({"function": "write_csv", "status": status})

        try:
            with open(target_name, "r", newline="\n") as reader:
                csvin = csv.DictReader(reader,
                                    skipinitialspace=True,
                                    delimiter=",",
                                    quotechar='"',
                                    quoting=csv.QUOTE_NONE)
                rows = {idx: values for idx, values in enumerate(csvin, 2)}

                for idx, value in data.items():
                    if value.get("remark") is not None:
                        if idx in self.change_rows.keys():
                            logging.info(f'"{value["remark"]}" Rows: "{idx}" in Target file\nUpdating records:"{self.change_rows[idx]}"')
                            value.popitem()
                            rows.update({idx: value})
                        elif idx in self.remove_rows:
                            continue
                    else:
                        rows[idx].update(data[idx])

            with open(target_name, "w", newline="\n") as writer:
                csvout = csv.DictWriter(writer,
                                        csvin.fieldnames,
                                        delimiter=",",
                                        quotechar='"',
                                        quoting=csv.QUOTE_NONE)
                csvout.writeheader()

                for idx in rows:
                    if idx not in self.remove_rows:
                        rows[idx].update({"CreateDate": rows[idx]["CreateDate"].strftime("%Y%m%d%H%M%S"),
                                        "LastLogin": rows[idx]["LastLogin"].strftime("%Y%m%d%H%M%S"),
                                        "LastUpdatedDate": rows[idx]["LastUpdatedDate"].strftime("%Y%m%d%H%M%S"),})
                        csvout.writerow(rows[idx])
            writer.closed

        except Exception as err:
            raise Exception(err)

        status = "succeed"
        self.logging[-1].update({"status": status})

        return status
