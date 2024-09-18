import logging
import re
from pathlib import Path
from os.path import join
import os
import glob
import pandas as pd
import numpy as np
import openpyxl
import chardet
from io import StringIO
from itertools import tee, chain
import xlrd
import csv
from .exception import CustomException
from .setup import Folder
from unicodedata import normalize

class Convert2File:
            
    async def check_source_file(self) -> None:
        
        logging.info('Check source file')

        for record in self.logging:
            record.update({'function': 'check_source_file'})
            
            if glob.glob(record['full_input'], recursive=True):
                status = 'found'
            else:
                status = 'not_found'
                record.update({'err': f'File not found {record['full_input']}'})

            record.update({'status': status})
            logging.info(f"Check source file: {record['full_input']}, for package: {record['package']}, status: {status}")

            if 'err' in record:
                raise CustomException(err=self.logging)
        
    async def separate_data_file(self) -> None:
        
        logging.info('Separate file from module')
        
        for i, record in enumerate(self.logging):
            record.update({'function': 'separate_data_from_file'})
            
            try:
                full_input = record['full_input']
                types = Path(full_input).suffix
                
                if ['.xlsx', '.xls'].__contains__(types):
                    format_file = self.read_excel_file(i, full_input)
                else:
                    format_file = self.read_file(i, full_input)
                    
                self.get_extract_file(i, format_file)
                        
            except Exception as err:
                record.update({'err': err})
            
            if 'err' in record:
                raise CustomException(err=self.logging)
            
            
    def read_excel_file(self, i: int, full_input: str) -> any:
        
        status = 'failed'
        self.logging[i].update({'function': 'read_excel_file', 'status': status})
        
        try:
            logging.info(f"Read format excel file: {full_input}, for package: {self.logging[i]['package']}")
            format_file = xlrd.open_workbook(full_input)
            
        except Exception as err:
            raise Exception(err)
        
        status = 'succeed'
        self.logging[i].update({'status': status})
        
        return format_file

    def read_file(self, i: int, full_input: str) -> any:

        status = 'failed'
        self.logging[i].update({'function': 'read_file', 'status': status})
        
        try:
            logging.info(f"Read format text/csv file: {full_input}, for run: {self.logging[i]['package']}")
            
            with open(full_input, 'rb') as f:
                file = f.read()
                encoding_result = chardet.detect(file)
            detected_encoding = encoding_result['encoding']
            content = file.decode(detected_encoding)
            
            match_encoding = re.findall(r'utf|ascii', detected_encoding, re.IGNORECASE)
            if match_encoding == []:
                # change encoding tis-620
                detected_encoding = 'tis-620'
                with open(full_input, 'r', encoding=detected_encoding) as f:
                    content = f.read()
            
            format_file = StringIO(content)
            
        except (LookupError, UnicodeDecodeError, TypeError) as err:
            raise Exception(err)
        
        status = 'succeed'
        self.logging[i].update({'status': status})
        
        return format_file
        
    def get_extract_file(self, i: int, format_file: any) -> dict:
        
        self.logging[i].update({'function': 'get_extract_data'})
        
        full_target = self.logging[i]['full_target']
        
        if re.search(r"PARAMLIST", full_target) is not None:
            columns = ['Parameter Name', 'Code values', 'Decode value']
            self.logging[i].update({'columns': columns})
            
            self.collect_param_file(i, format_file)
        else:
            columns = ['ApplicationCode', 'AccountOwner', 'AccountName', 'AccountType', 'EntitlementName', 'SecondEntitlementName', 'ThirdEntitlementName','AccountStatus', 
                        'IsPrivileged', 'AccountDescription', 'CreateDate', 'LastLogin', 'LastUpdatedDate', 'AdditionalAttribute', 'Country']
            self.logging[i].update({'columns': columns})
            
            self.collect_user_file(i, format_file)
        
    def comparing_dataframes(self, i: int, df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
        
        # logging.info('Comparing two dataframes and getting the differences')
        
        status = "failed"
        self.logging[i].update({'function': 'compare_data', 'status': status})
        
        def remark_rows(df):
            if set('remark').issubset(df.columns):    
                df = df.loc[df['remark'] != 'Remove']
            else:
                df['remark'] = 'Insert'
            return df
        
        try:
            self.merge_index = np.union1d(df.index, new_df.index)
            
            df = remark_rows(df)
            df = df.reindex(index=self.merge_index, columns=df.columns).iloc[:,:-1]
            
            new_df = remark_rows(new_df)
            self.new_df = new_df.reindex(index=self.merge_index, columns=new_df.columns).iloc[:,:-1]
            
            df['count'] = pd.DataFrame(np.where(df.ne(self.new_df), True, df), index=df.index, columns=df.columns)\
                .apply(lambda x: (x == True).sum(), axis=1)
            
        except Exception as err:
            raise Exception(err)
        
        status = 'succeed'
        self.logging[i].update({'status': status})
        
        return df
    
    def change_data_capture(self, i: int, df: pd.DataFrame) -> dict:
        
        # logging.info('Change data capture')
        
        status = 'failed'
        self.logging[i].update({'function': 'data_change_capture', 'status': status})
        
        self.update_rows = {}
        def format_record(record):
            return "\n".join("{!r}: {!r};".format(columns, values) for columns, values in record.items())
        
        self.remove_rows = []
        if len(df.index) > len(self.new_df.index):
            self.remove_rows = [idx for idx in list(df.index) if idx not in list(self.new_df.index)]
            
        try:
            i = 0
            for idx, row in enumerate(self.merge_index, 2):
                if row not in self.remove_rows:
                    record = {}
                    for data, change_data in zip(df.items(), self.new_df.items()):
                        ## No Change
                        if df.loc[row, 'count'] not in [3, 15]:
                            if df.loc[row, 'count'] < 1:
                                df.loc[row, data[0]] = data[1][row]
                                df.loc[row, 'remark'] = 'No_change'
                            else:
                                ## Update
                                if data[1][row] != change_data[1][row]:
                                    record.update({data[0]: f"{data[1][row]} => {change_data[1][row]}"})
                                df.loc[row, data[0]] = change_data[1][row]
                                df.loc[row, 'remark'] = 'Update'
                        else:
                            ## Insert
                            record.update({data[0]: change_data[1][row]})
                            df.loc[row, data[0]] = change_data[1][row]
                            df.loc[row, 'remark'] = 'Insert'
                            
                    if record != {}:
                        self.update_rows[idx] = format_record(record)
                else:
                    ## Remove
                    self.remove_rows[i] = idx
                    df.loc[row, 'remark'] = 'Remove'
                    i += 1
            
            df = df.drop(['count'], axis=1)
            rows = 2
            df.index += rows
            cdc = df.to_dict(orient='index')
            
        except Exception as err:
            raise Exception(err)
        
        status = "succeed"
        self.logging[i].update({'status': status})
        
        return cdc

    async def genarate_tmp_file(self) -> None:
        
        self.clear_tmp_file()
        
        logging.info('Genarate data to tmp file')
        
        status = "failed"
        for i, record in enumerate(self.logging):
            try:
                ## Read tmp file
                tmp_dir = join(Folder.TMP, self.module, self.date.strftime('%Y%m%d'))
                os.makedirs(tmp_dir, exist_ok=True)
                tmp_name = f"{Path(record['full_target']).stem}.xlsx"
                full_tmp = join(tmp_dir, tmp_name)
                record.update({"full_tmp": full_tmp})
                
                ## Set dataframe from tmp file 
                self.create_workbook(i)            
                data = self.sheet.values
                columns = next(data)[0:]
                tmp_df = pd.DataFrame(data, columns=columns).replace([None], [''])
                
                ## Set dataframe from raw file      
                raw_df = pd.DataFrame(record['data']).replace([None],['NA'])
                
                ## Validate data change row by row
                cmp_df = self.comparing_dataframes(i, tmp_df, raw_df)
                cdc = self.change_data_capture(i, cmp_df)
                
                ## Write tmp file
                status = self.write_worksheet(i, cdc)
                record.update({'function': 'genarate_tmp_file', 'status': status})
                logging.info(f"Write data to tmp file: {record['full_tmp']}, status: {status}")
                
            except Exception as err:
                record.update({'err': err})

            if 'err' in record:
                raise CustomException(err=self.logging)
            
    def create_workbook(self, i:int) -> None:
        
        logging.info(f"Create tmp file: {self.logging[i]['full_tmp']}")
        
        status = 'failed'
        self.logging[i].update({'function': 'create_workbook', 'status': status})

        try:
            self.create = False
            self.workbook = openpyxl.load_workbook(self.logging[i]['full_tmp'])
            get_sheet = self.workbook.get_sheet_names()
            self.sheet_num = len(get_sheet)
            self.sheet_name = f"RUN_TIME_{self.sheet_num}"
            
            if self.sheet_name in get_sheet:
                self.create = True
                self.sheet = self.workbook.get_sheet_by_name(self.sheet_name)

        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.sheet_name = 'RUN_TIME_1'
            self.sheet =  self.workbook.get_sheet_by_name('Sheet')
            self.sheet.title = self.sheet_name
            self.sheet.append(self.logging[i]['columns'])
            
            self.workbook.save(self.logging[i]['full_tmp'])
            self.sheet_num = 1
            
        status = 'succeed'
        self.logging[i].update({'status': status})
        
    def write_worksheet(self, i: int, cdc: dict) -> str:
        
        status = 'failed'
        if self.create:
            self.sheet_name = f"RUN_TIME_{self.sheet_num + 1}"
            self.sheet = self.workbook.create_sheet(self.sheet_name)

        logging.info(f"Write to sheet: {self.sheet_name}")
        
        rows = 2
        max_row = max(cdc, default=0)
        self.logging[i].update({'function': 'write_worksheet', 'sheet_name': self.sheet_name, 'status': status,})
        
        try:
            # write column
            for idx, col in enumerate(cdc[rows].keys(), 1):
                self.sheet.cell(row=1, column=idx).value = col
            ## write row
            while rows <= max_row:
                for idx, col in enumerate(cdc[rows].keys(), 1):
                    self.sheet.cell(row=rows, column=idx).value = cdc[rows][col]
                    if col == 'remark':
                        if rows in self.remove_rows:
                            ## remove row
                            record = f"{cdc[rows][col]} rows: ({rows}) in tmp file"
                        elif rows in self.update_rows.keys():
                            ## update / insert row
                            record = f"{cdc[rows][col]} rows: ({rows}) in tmp file, Updating records: ({self.update_rows[rows]})"
                        else:
                            continue
                        #logging.info(record)
                rows += 1
                
        except KeyError as err:
            raise KeyError(f"Can not write rows: {err} in tmp file")
        
        ## save file
        self.sheet.title = self.sheet_name
        self.workbook.active = self.sheet
        self.workbook.move_sheet(self.workbook.active, offset=-self.sheet_num)
        self.workbook.save(self.logging[i]['full_tmp'])

        status = 'succeed'
        self.logging[i].update({"status": status})
        
        return status

    async def genarate_target_file(self) -> None:

        logging.info('Genarate data to target file')

        status = 'failed'
        for i, record in enumerate(self.logging):
            try:
                ## Set dataframe from tmp/raw file
                if self.store_tmp is True:
                    workbook = openpyxl.load_workbook(record['full_tmp'])
                    sheet = workbook.get_sheet_by_name(record['sheet_name'])
                    data = sheet.values
                    columns = next(data)[0:]
                    new_df = pd.DataFrame(data, columns=columns).replace([None], [''])
                else:
                    new_df = pd.DataFrame(record['data']).replace([None],['NA'])
                    
                ## Set dataframe from target file
                try:
                    target_df = self.read_csv_file(i, record['full_target'])
                    
                except FileNotFoundError:
                    target_df = pd.DataFrame(columns=record['columns'])
                    target_df.to_csv(record["full_target"], index=None, header=True, sep=",")
                
                # Validate data change row by row
                cmp_df = self.comparing_dataframes(i, target_df, new_df)
                cdc = self.change_data_capture(i, cmp_df)
                
                ## Write csv file
                status = self.write_csv(i, cdc)
                record.update({'function': 'genarate_target_file', 'status': status})
                logging.info(f"Write data to target file: {record['full_target']}, status: {status}")

            except Exception as err:
                record.update({'err': err})

        if 'err' in record:
            raise CustomException(err=self.logging)
        
    def read_csv_file(self, i: int, file: str) -> pd.DataFrame:
        
        logging.info(f'Read csv file: {file}')
        
        status = 'failed'
        self.logging[i].update({'function': 'read_csv', 'status': status})
        
        data = []
        with open(file, 'r', newline="\n", encoding='utf-8') as reader:
            csvin = csv.reader(
                reader,
                skipinitialspace=True,
                delimiter=",",
                quotechar='"',
                quoting=csv.QUOTE_ALL)
            header = next(csvin)
                
            ## skip last row (total row)
            last_row, row = tee(csvin)
            next(chain(last_row, range(1)))
            for _ in last_row:
                data.append(next(row))
                    
        df = pd.DataFrame(data, columns=header)
        
        status = 'succeed'
        self.logging[i].update({'status': status})
        
        return df
    
    def write_csv(self, i: int, cdc: dict) -> str:
        
        # logging.info(f'Write mode: {self.write_mode}')

        status = 'failed'
        self.logging[i].update({'function': 'write_csv', 'status': status})

        try:
            with open(self.logging[i]['full_target'], 'r', newline="\n", encoding='utf-8') as reader:
                csvin = csv.DictReader(
                    reader,
                    skipinitialspace=True,
                    delimiter=',',
                    quotechar='"',
                    quoting=csv.QUOTE_ALL,)
                
                ## skip last row (total row)
                last_row, row = tee(csvin)
                next(chain(last_row, range(1)))
            
                rows = {}
                for idx , _ in enumerate(last_row, 2):
                    rows.update({idx: next(row)})
                
                for idx, value in cdc.items():
                    if value.get('remark') is not None:
                        if idx in self.update_rows.keys():
                            # logging.info(f"{value['remark']} rows: ({idx}) in target file, Updating records: ({self.update_rows[idx]})")
                            value.popitem()
                            rows.update({idx: value})
                            
                        elif idx in self.remove_rows:
                            continue
                    else:
                        rows[idx] = cdc[idx]
            
            with open(self.logging[i]['full_target'], 'w', newline="\n", encoding='utf-8') as writer:
                csvout = csv.DictWriter(
                    writer,
                    csvin.fieldnames,
                    delimiter=',',
                    quotechar='"',
                    quoting=csv.QUOTE_ALL,)
                csvout.writeheader()
            
                for idx in rows:
                    if idx not in self.remove_rows:
                        csvout.writerow(rows[idx])
                writer.close()
            
            with open(self.logging[i]['full_target'], mode='a', newline='\n', encoding='utf-8') as writer:
                writer.write('"{}","{}"'.format('TotalCount', len(rows)))
                logging.info(f'Total count: ({len(rows)}) row')
                writer.close()
            
        except Exception as err:
            raise Exception(err)

        status = 'succeed'
        self.logging[i].update({'status': status})
        return status
