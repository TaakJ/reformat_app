import logging
from pathlib import Path
from os.path import join
import os
import glob
import shutil
import pandas as pd
import numpy as np
import re
import openpyxl
import chardet
from io import StringIO
from itertools import tee, chain
import xlrd
import csv
from .exception import CustomException
from .setup import Folder

class Convert2File:

    async def check_source_file(self) -> None:
        
        logging.info("Check source file")
        
        for record in self.logging:
            record.update({"function": "check_source_file"})
            full_input = record["full_input"]
            
            if glob.glob(full_input, recursive=True):
                status = "found"
            else:
                status = "not_found"
                record.update({"err": f"File not found {full_input}"})
            
            record.update({"status": status})
            logging.info(f"Check source file: {full_input}, status: {status}")
                
            if "err" in record:
                raise CustomException(err=self.logging)
        
    async def separate_data_file(self) -> None:
        
        logging.info("Separate file from module")
        
        for i, record in enumerate(self.logging):
            record.update({"function": "separate_data_from_file"})
            
            try:
                types = Path(record["full_input"]).suffix
                status_file = record["status"]
                
                if status_file == "found":
                    if [".xlsx", ".xls"].__contains__(types):
                        self.read_excel_file(i)
                    else:
                        self.read_file(i)
                else:
                    continue

            except Exception as err:
                record.update({"err": err})
            
            if "err" in record:
                raise CustomException(err=self.logging)
            
    def read_excel_file(self, i: int) -> any:

        status = "failed"
        self.logging[i].update({"function": "read_excel_file", "status": status})
        
        try:
            full_input = self.logging[i]["full_input"]
            workbook = xlrd.open_workbook(full_input)
            
            data = self.get_extract_data(i, workbook)

        except Exception as err:
            raise Exception(err)
        
        status = "succeed"
        self.logging[i].update({"status": status})
        
        return data

    def read_file(self, i: int) -> any:

        status = "failed"
        self.logging[i].update({"function": "read_file", "status": status})
        
        try:
            full_input = self.logging[i]["full_input"]
            logging.info(f"Read format text/csv file: {full_input}")
            
            with open(full_input, 'rb') as f:
                file = f.read()
            encoding_result = chardet.detect(file)
            encoding = encoding_result['encoding']
            line = StringIO(file.decode(encoding))
            
            ## call function collect data in module
            self.get_extract_data(i, line)
            
        except Exception as err:
            raise Exception(err)
        
        status = "succeed"
        self.logging[i].update({"status": status})
        
    def set_initial_data_type(self, i: int, df: pd.DataFrame) -> pd.DataFrame:
        
        self.logging[i].update({"function": "set_initial_data_type"})
        full_input = self.logging[i]["full_input"]
        
        if re.search(r'Param', full_input) is not None:
            template_name = "Param Requirements.xlsx"
            df = self.param_type(df)
        else:
            template_name = "Application Data Requirements.xlsx"
            df = self.application_type(df)
            
        self.logging[i].update({"template": template_name})
        return df
        
    def application_type(self, df: pd.DataFrame) -> pd.DataFrame:
        try:
            df = df.apply(lambda x: x.str.strip()).replace('', 'NA')
            df = df.astype({
                    "ApplicationCode": object,
                    "AccountOwner": object,
                    "AccountName": object,
                    "AccountType": object,
                    "EntitlementName": object,
                    "SecondEntitlementName": object,
                    "ThirdEntitlementName": object,
                    "AccountStatus": object,
                    "IsPrivileged": object,
                    "AccountDescription": object,
                    "CreateDate": object,
                    "LastLogin": object,
                    "LastUpdatedDate": object,
                    "AdditionalAttribute": object,})
            if "remark" in df.columns:
                df = df.loc[df["remark"] != "Remove"]
            else:
                df["remark"] = "Insert"

        except Exception as err:
            raise Exception(err)
        return df
    
    def param_type(self, df: pd.DataFrame) -> pd.DataFrame:
        try:
            df = df.apply(lambda x: x.str.strip()).replace('', 'NA')
            df = df.astype({
                    "Parameter Name": object,
                    "Code value": object,
                    "Decode value": object,})
            if "remark" in df.columns:
                df = df.loc[df["remark"] != "Remove"]
            else:
                df["remark"] = "Insert"

        except Exception as err:
            raise Exception(err)
        return df

    async def genarate_tmp_file(self) -> None:
        
        self.clear_tmp()
        
        logging.info("Genarate data to tmp file")

        status = "failed"
        for i, record in enumerate(self.logging):
            try:            
                ## Read tmp file
                tmp_dir = join(Folder.TMP, self.module, self.date.strftime("%Y%m%d"))
                os.makedirs(tmp_dir, exist_ok=True)
                tmp_name = f"T_{Path(record["full_target"]).stem}.xlsx"
                full_tmp = join(tmp_dir, tmp_name)
                record.update({"full_tmp": full_tmp})
                
                ## Set dataframe from tmp file 
                self.create_workbook(i)            
                data = self.sheet.values
                columns = next(data)[0:]
                tmp_df = pd.DataFrame(data, columns=columns)
                tmp_df = self.set_initial_data_type(i, tmp_df)
                
                ## Set dataframe from raw file      
                raw_df = pd.DataFrame(record["data"])
                
                ## Validate data change row by row
                cmp_df = self.comparing_dataframes(i, tmp_df, raw_df)
                cdc = self.change_data_capture(i, cmp_df)
                        
                ## Write tmp file
                status = self.write_worksheet(i, cdc)

            except Exception as err:
                record.update({"err": err})
                
            record.update({"function": "genarate_tmp_file", "status": status})
            logging.info(f"Write data to tmp file: {record["full_tmp"]}, status: {status}")

            if "err" in record:
                raise CustomException(err=self.logging)
            
    def create_workbook(self, i:int) -> None:

        full_tmp = self.logging[i]["full_tmp"]
        logging.info(f"Create tmp file: {full_tmp}")

        status = "failed"
        self.logging[i].update({"function": "create_workbook", "status": status})

        try:
            self.create = False
            self.workbook = openpyxl.load_workbook(full_tmp)
            get_sheet = self.workbook.get_sheet_names()
            self.sheet_num = len(get_sheet)
            self.sheet_name = f"RUN_TIME_{self.sheet_num - 1}"

            if self.sheet_name in get_sheet:
                self.create = True
                self.sheet = self.workbook.get_sheet_by_name(self.sheet_name)
            else:
                self.sheet = self.workbook.get_sheet_by_name("Field Name")

        except FileNotFoundError:
            template_name = self.logging[i]["template"]
            full_template = join(Folder.TEMPLATE, template_name)
            try:
                if not glob.glob(full_tmp, recursive=True):
                    shutil.copy2(full_template, full_tmp)
            except:
                raise
            
            self.workbook = openpyxl.load_workbook(full_tmp)
            self.sheet = self.workbook.get_sheet_by_name("Field Name")
            self.sheet_name = "RUN_TIME_1"
            self.sheet_num = 1

        status = "succeed"
        self.logging[i].update({"status": status})

    def comparing_dataframes(self, i: int, df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
        
        logging.info("Comparing two dataframes and getting the differences")
        
        status = "failed"
        self.logging[i].update({"function": "compare_data", "status": status})
        
        try:
            ## Merge index
            self.merge_index = np.union1d(df.index, new_df.index)
            ## As starter dataframe for compare
            df = df.reindex(index=self.merge_index, columns=df.columns).iloc[:,:-1]
            ## Change data / new data
            self.new_df = new_df.reindex(index=self.merge_index, columns=new_df.columns).iloc[:,:-1]
            ## Compare data
            df["count"] = pd.DataFrame(np.where(df.ne(self.new_df), True, df), index=df.index, columns=df.columns)\
                .apply(lambda x: (x == True).sum(), axis=1)
            
        except Exception as err:
            raise Exception(err)
        
        status = "succeed"
        self.logging[i].update({"status": status})
        return df
    
    def change_data_capture(self, i: int, df: pd.DataFrame) -> dict:
        
        logging.info("Change data capture")
        
        status = "failed"
        self.logging[i].update({"function": "data_change_capture", "status": status})
        
        def format_record(record):
            return "\n".join("{!r}: {!r};".format(columns, values) for columns, values in record.items())
        
        self.update_rows = {}
        self.remove_rows = []
        if len(df.index) > len(self.new_df.index):
            self.remove_rows = [idx for idx in list(df.index) if idx not in list(self.new_df.index)]
            
        try:
            i = 0
            for idx, row in enumerate(self.merge_index, 2):
                if row not in self.remove_rows:
                    record = {}
                    for data, change_data in zip(df.items(), self.new_df.items()):
                        ## No Change
                        if df.loc[row, "count"] not in [3, 15]:
                            if df.loc[row, "count"] < 1:
                                df.loc[row, data[0]] = data[1][row]
                                df.loc[row, "remark"] = "No_change"
                            else:
                                ## Update
                                if data[1][row] != change_data[1][row]:
                                    record.update({data[0]: f"{data[1][row]} => {change_data[1][row]}"})
                                df.loc[row, data[0]] = change_data[1][row]
                                df.loc[row, "remark"] = "Update"
                        else:
                            ## Insert
                            record.update({data[0]: change_data[1][row]})
                            df.loc[row, data[0]] = change_data[1][row]
                            df.loc[row, "remark"] = "Insert"
                            
                    if record != {}:
                        self.update_rows[idx] = format_record(record)
                else:
                    ## Remove
                    self.remove_rows[i] = idx
                    df.loc[row, "remark"] = "Remove"
                    i += 1
            
            df = df.drop(["count"], axis=1)
            rows = 2
            df.index += rows
            cdc = df.to_dict(orient="index")
            
        except Exception as err:
            raise Exception(err)
        
        status = "succeed"
        self.logging[i].update({"status": status})
        return cdc

    def write_worksheet(self, i: int, cdc: dict) -> str:
        
        status = "failed"
        if self.create:
            self.sheet_name = f"RUN_TIME_{self.sheet_num}"
            self.sheet = self.workbook.create_sheet(self.sheet_name)

        logging.info(f"Write to sheet: {self.sheet_name}")

        rows = 2
        max_row = max(cdc, default=0)
        self.logging[i].update({"function": "write_worksheet", "sheet_name": self.sheet_name, "status": status,})
        
        try:
            # write column
            for idx, col in enumerate(cdc[rows].keys(), 1):
                self.sheet.cell(row=1, column=idx).value = col

            ## write row
            while rows <= max_row:
                for idx, col in enumerate(cdc[rows].keys(), 1):
                    if col == "remark":
                        if rows in self.remove_rows:
                            ## remove row
                            write_row = (f"{cdc[rows][col]} rows: ({rows}) in tmp file")
                        elif rows in self.update_rows.keys():
                            ## update / insert row
                            write_row = f"{cdc[rows][col]} rows: ({rows}) in tmp file, Updating records: ({self.update_rows[rows]})"
                        # else:
                        #     ## no change row
                        #     write_row = f"No change rows: ({rows}) in tmp file" ## No change row
                        logging.info(write_row)
                            
                    self.sheet.cell(row=rows, column=idx).value = cdc[rows][col]
                rows += 1
                
        except KeyError as err:
            raise KeyError(f"Can not write rows: {err} in tmp file")

        ## save file
        full_tmp = self.logging[i]["full_tmp"]
        self.sheet.title = self.sheet_name
        self.workbook.active = self.sheet
        self.workbook.move_sheet(self.workbook.active, offset=-self.sheet_num)
        self.workbook.save(full_tmp)

        status = "succeed"
        self.logging[i].update({"status": status})
        return status

    async def genarate_target_file(self) -> None:

        logging.info("Genarate data to target file")

        status = "failed"
        for i, record in enumerate(self.logging):
            try:
                ## Set dataframe from tmp/raw file
                if self.store_tmp is True:
                    workbook = openpyxl.load_workbook(record["full_tmp"])
                    sheet = workbook.get_sheet_by_name(record["sheet_name"])
                    data = sheet.values
                    columns = next(data)[0:]
                    new_df = pd.DataFrame(data, columns=columns)
                else:
                    new_df = pd.DataFrame(record["data"])
                new_df = self.set_initial_data_type(i, new_df)
                
                ## Set dataframe from target file
                try:
                    target_df = self.read_csv_file(i)
                    
                except FileNotFoundError:
                    template_name = join(Folder.TEMPLATE, record["template"])
                    target_df = pd.read_excel(template_name)
                    target_df.to_csv(record["full_target"], index=None, header=True, sep=",")
                target_df = self.set_initial_data_type(i, target_df)
                
                # Validate data change row by row
                cmp_df = self.comparing_dataframes(i, target_df, new_df)
                cdc = self.change_data_capture(i, cmp_df)
                
                ## Write csv file
                status = self.write_csv(i, cdc)

            except Exception as err:
                record.update({"err": err})
                
            record.update({"function": "genarate_target_file", "state": status})
            logging.info(f"Write data to target file: {record["full_target"]}, status: {status}")

        if "err" in record:
            raise CustomException(err=self.logging)
        
    def read_csv_file(self, i: int) -> pd.DataFrame:
        
        full_target = self.logging[i]["full_target"]
        logging.info(f"Read csv file: {full_target}")
        
        status = "failed"
        self.logging[i].update({"function": "read_csv", "status": status})
        
        data = []
        with open(full_target, "r", newline="\n") as reader:
            csvin = csv.reader(
                reader,
                skipinitialspace=True,
                delimiter=",",
                quotechar='"',
                quoting=csv.QUOTE_ALL)
            header = next(csvin)
                
            ## skip last row (total row)
            last_row, row = tee(csvin)
            next(chain(last_row, range(1)))
            for _ in last_row:
                data.append(next(row))
                    
        df = pd.DataFrame(data, columns=header)
        
        status = "succeed"
        self.logging[i].update({"status": status})
        return df
    
    def write_csv(self, i: int, cdc: dict) -> str:
        
        full_target = self.logging[i]["full_target"]
        logging.info(f"Write mode: {self.write_mode}")

        status = "failed"
        self.logging[i].update({"function": "write_csv", "status": status})

        try:
            with open(full_target, "r", newline="\n") as reader:
                csvin = csv.DictReader(
                    reader,
                    skipinitialspace=True,
                    delimiter=",",
                    quotechar='"',
                    quoting=csv.QUOTE_ALL,)
                
                ## skip last row (total row)
                last_row, row = tee(csvin)
                next(chain(last_row, range(1)))
            
                rows = {}
                for idx , _ in enumerate(last_row, 2):
                    rows.update({idx: next(row)})
                
                for idx, value in cdc.items():
                    if value.get("remark") is not None:
                        if idx in self.update_rows.keys():
                            logging.info(f'{value["remark"]} rows: ({idx}) in target file, Updating records: ({self.update_rows[idx]})')
                            value.popitem()
                            rows.update({idx: value})
                            
                        elif idx in self.remove_rows:
                            continue
                    else:
                        rows[idx] = cdc[idx]
            
            with open(full_target, "w", newline="\n") as writer:
                csvout = csv.DictWriter(
                    writer,
                    csvin.fieldnames,
                    delimiter=",",
                    quotechar='"',
                    quoting=csv.QUOTE_ALL,)
                csvout.writeheader()
            
                for idx in rows:
                    if idx not in self.remove_rows:
                        csvout.writerow(rows[idx])
                writer.close()
            
            with open(full_target, mode="a", newline="\n") as writer:
                writer.write('"{}","{}"'.format("TotalCount", len(rows)))
                logging.info(f'Total count: ({len(rows)}) row')
                writer.close()
            
        except Exception as err:
            raise Exception(err)

        status = "succeed"
        self.logging[i].update({"status": status})
        return status
