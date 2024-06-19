from exception import CustomException
from setup import Folder
import logging
from pathlib import Path
from os.path import join
import glob
import shutil
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import chardet
from io import StringIO
import re
import xlrd
import csv

class convert_2_files:

    async def check_source_files(self) -> None:

        logging.info("Check Source files..")

        set_log = []
        for _dir in self.input_dir:

            status_file = "not_found"
            if glob.glob(_dir, recursive=True):
                status_file = "found"

            record = {
                "module": self.module,
                "input_dir": _dir,
                "status_file": status_file,
                "function": "check_source_files",
            }
            set_log.append(record)
            logging.info(f'Source file: "{_dir}", Status: "{status_file}"')

        self._log_setter(set_log)

    async def retrieve_data_from_source_files(self) -> None:

        logging.info("Retrieve Data from Source files..")

        state = "failed"
        for i, record in enumerate(self.logging):
            record.update({"function": "retrieve_data_from_source_files", "state": state})

            _dir = record["input_dir"]
            types = Path(_dir).suffix
            status_file = record["status_file"]
            try:
                if status_file == "found":
                    if [".xlsx", ".xls"].__contains__(types):
                        logging.info(f'Read Excel file: "{_dir}"')
                        _data = self.excel_data_cleaning(i)

                    else:
                        logging.info(f'Read Text file: "{_dir}"')
                        _data = self.text_data_cleaning(i)
                else:
                    continue
                state = "succeed"
                record.update({"data": _data, "state": state})

            except Exception as err:
                record.update({"errors": err})

            if "errors" in record:
                raise CustomException(errors=self.logging)

    def read_text_files(func):
        def wrapper(*args: tuple, **kwargs: dict) -> dict:

            by_lines = iter(func(*args, **kwargs))
            _data = {}

            rows = 0
            while True:
                try:
                    list_by_lines = []
                    for sheets, data in next(by_lines).items():

                        if sheets == "LDS":
                            if rows == 0:
                                ## herder column
                                list_by_lines = " ".join(data).split(" ")
                            else:
                                ## row value
                                for idx, value in enumerate(data):
                                    if idx == 0:
                                        value = re.sub(r"\s+", ",", value).split(",")
                                        list_by_lines.extend(value)
                                    else:
                                        list_by_lines.append(value)

                        elif sheets == "DOC":
                            if rows == 1:
                                ## herder column
                                list_by_lines = " ".join(data).split(" ")
                            elif rows > 1:
                                ## row value
                                for idx, value in enumerate(data):
                                    if idx == 3:
                                        value = re.sub(r"\s+", ",", value).split(",")
                                        list_by_lines.extend(value)
                                    else:
                                        list_by_lines.append(value)

                        elif sheets == "ADM":
                            ## row value
                            list_by_lines = data

                        if list_by_lines != []:
                            if sheets not in _data:
                                _data[sheets] = [list_by_lines]
                            else:
                                _data[sheets].append(list_by_lines)
                        else:
                            continue
                    rows += 1

                except StopIteration:
                    break
            return _data

        return wrapper

    @read_text_files
    def text_data_cleaning(self, i: int) -> any:

        # logging.info("Cleansing Data From Text file..")
        self.logging[i].update({"function": "text_data_cleaning"})

        _dir = self.logging[i]["input_dir"]
        sheets = self.logging[i]["module"]

        files = open(_dir, "rb")
        encoded = chardet.detect(files.read())["encoding"]
        files.seek(0)
        decode_data = StringIO(files.read().decode(encoded))

        for lines in decode_data:
            regex = re.compile(r"\w+.*")
            find_regex = regex.findall(lines)
            if find_regex != []:
                yield {sheets: re.sub(r"\W\s+","||","".join(find_regex).strip()).split("||")}

    def read_excle_files(func):
        def wrapper(*args: tuple, **kwargs: dict) -> dict:

            by_sheets = iter(func(*args, **kwargs))
            _data = {}

            while True:
                try:
                    for sheets, data in next(by_sheets).items():
                        if not all(dup == data[0] for dup in data) and \
                            not data.__contains__("Centralized User Management : User List."):
                            if sheets not in _data:
                                _data[sheets] = [data]
                            else:
                                _data[sheets].append(data)

                except StopIteration:
                    break
            return _data

        return wrapper

    @read_excle_files
    def excel_data_cleaning(self, i: int) -> any:

        # logging.info("Cleansing Data From Excle files..")
        self.logging[i].update({"function": "excel_data_cleaning"})

        workbook = xlrd.open_workbook(self.logging[i]["input_dir"])
        sheet_list = [sheet for sheet in workbook.sheet_names() if sheet != "StyleSheet"]

        for sheets in sheet_list:
            cells = workbook.sheet_by_name(sheets)
            for row in range(0, cells.nrows):
                yield {sheets: [cells.cell(row, col).value for col in range(cells.ncols)]}

    async def write_data_to_tmp_file(self) -> None:

        logging.info("Write Data to Tmp files..")

        state = "failed"
        for record in self.logging:
            try:
                if record["module"] == "Target_file":

                    tmp_name = join(Folder.TMP,f"TMP_{self.module}-{self.batch_date.strftime('%d%m%y')}.xlsx")
                    record.update({"input_dir": tmp_name,"function": "write_data_to_tmp_file","state": state})

                    # return value from mapping column
                    change_df = pd.DataFrame(record["data"])
                    change_df["remark"] = "Inserted"
                    change_df[["CreateDate","LastUpdatedDate"]] = change_df[["CreateDate","LastUpdatedDate"]]\
                        .apply(pd.to_datetime, format="%Y%m%d%H%M%S")

                    try:
                        workbook = openpyxl.load_workbook(tmp_name)
                        get_sheet = workbook.get_sheet_names()
                        sheet_num = len(get_sheet)
                        sheet_name = f"RUN_TIME_{sheet_num - 1}"
                        sheet = workbook.get_sheet_by_name(sheet_name)
                        workbook.active = sheet_num

                    except FileNotFoundError:
                        template_name = join(Folder.TEMPLATE, "Application Data Requirements.xlsx")
                        try:
                            if not glob.glob(tmp_name, recursive=True):
                                state = shutil.copy2(template_name, tmp_name)
                            state = "succeed"
                        except:
                            raise
                        workbook = openpyxl.load_workbook(tmp_name)
                        sheet = workbook.worksheets[0]
                        sheet_name = "RUN_TIME_1"
                        sheet_num = 1
                        sheet.title = sheet_name

                    logging.info(f"Generate Sheet_name: {sheet_name} in Tmp files.")

                    ## read tmp files.
                    data = sheet.values
                    columns = next(data)[0:]
                    tmp_df = pd.DataFrame(data, columns=columns)
                    tmp_df["remark"] = "Inserted"
                    tmp_df[["CreateDate","LastUpdatedDate"]] = tmp_df[["CreateDate","LastUpdatedDate"]]\
                        .apply(pd.to_datetime, format="%Y%m%d%H%M%S")
                        
                    if state != "succeed":
                        tmp_df = tmp_df.loc[tmp_df["remark"] != "Removed"]
                        sheet_name = f"RUN_TIME_{sheet_num}"
                        sheet = workbook.create_sheet(sheet_name)
                    
                    _data = self.validation_data(tmp_df, change_df)
                    
                    ## write to tmp files.
                    state = self.write_worksheet(sheet, _data)
                    workbook.move_sheet(workbook.active, offset=-sheet_num)
                    workbook.save(tmp_name)

                    record.update({"sheet_name": sheet_name, "state": state})
                    logging.info(f"Write to Tmp files state: {state}.")

            except Exception as err:
                record.update({"errors": err})

            if "errors" in record:
                raise CustomException(errors=self.logging)

    def write_worksheet(self, sheet: any, change_data: dict) -> str:

        self.logging[-1].update({"function": "write_worksheet"})
        max_rows = max(change_data, default=0)
        logging.info(f"Data for write: {max_rows}. rows")

        start_rows = 2
        try:
            # write columns.
            for idx, columns in enumerate(change_data[start_rows].keys(), 1):
                sheet.cell(row=1, column=idx).value = columns
                
            ## write data.
            while start_rows <= max_rows:
                for idx, columns in enumerate(change_data[start_rows].keys(),1):
                    if columns == "remark":
                        if start_rows in self.remove_rows and change_data[start_rows][columns] == "Removed":
                            ## Removed data.
                            _show = f"{change_data[start_rows][columns]} Rows: ({start_rows}) in Tmp files."
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                            
                        elif start_rows in self.change_rows.keys() and change_data[start_rows][columns] in ["Inserted","Updated"]:
                            ## Updated / Insert data. 
                            _show = f"{change_data[start_rows][columns]} Rows: ({start_rows}) in Tmp files.\
                                Record Changed: {self.change_rows[start_rows]}"
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                        else:
                            ## No change data.
                            _show = f"No Change Rows: ({start_rows}) in Tmp files."
                            sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                            
                        logging.info(_show)
                        
                    elif columns in ["CreateDate","LastUpdatedDate"]:
                        sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns].strftime("%Y%m%d%H%M%S")
                        
                    else:
                        sheet.cell(row=start_rows, column=idx).value = change_data[start_rows][columns]
                        
                start_rows += 1
            state = "succeed"
            
        except KeyError as err:
            raise KeyError(f"Can not Write rows: {err} in Tmp files.")
        return state

    async def write_data_to_target_file(self) -> None:

        logging.info("Write Data to Target files..")

        state = "failed"
        for record in self.logging:
            try:
                
                if record["module"] == "Target_file":
                    record.update({"function": "write_data_to_target_file", "state": state})
                    
                    ## read tmp file / return value from mapping column
                    if self.store_tmp is True:
                        tmp_name = record["input_dir"]
                        sheet_name = record["sheet_name"]
                        change_df = pd.read_excel(tmp_name, sheet_name=sheet_name,)
                        change_df = change_df.loc[change_df['remark'] != "Removed"]
                    else:
                        change_df = pd.DataFrame(record["data"])
                        change_df["remark"] = "Inserted"
                    change_df[["CreateDate","LastUpdatedDate"]] = change_df[["CreateDate","LastUpdatedDate"]]\
                        .apply(pd.to_datetime, format="%Y%m%d%H%M%S")
                    
                    ## read target file (csv).
                    if self.write_mode == "overwrite" or self.manual:
                        target_name = join(self.output_dir, self.output_file)
                    else:
                        suffix = f"{self.batch_date.strftime('%d%m%y')}"
                        self.output_file = f"{Path(self.output_file).stem}_{suffix}.csv"
                        target_name = join(self.output_dir, self.output_file)
                        
                    record.update({"input_dir": target_name})
                    
                    try:
                        target_df = pd.read_csv(target_name)
                        target_df["remark"] = "Inserted"
                        target_df[["CreateDate","LastUpdatedDate"]] = target_df[["CreateDate","LastUpdatedDate"]]\
                            .apply(pd.to_datetime, format="%Y%m%d%H%M%S")
                        
                    except FileNotFoundError:
                        template_name = join(Folder.TEMPLATE, "Application Data Requirements.xlsx")
                        target_df = pd.read_excel(template_name)
                        target_df.to_csv(target_name, index=None, header=True)
                    
                    data = self.customize_data(target_df, change_df)
                    record.update({"function": "write_csv", "state": state})
                    
                    ## write csv file
                    state = self.write_csv(target_name, data)
                    record.update({"state": state})
                    
                    logging.info(f"Write to Target Files status: {state}.")
            
            except Exception as err:
                record.update({"errors": err})

        if "errors" in record:
            raise CustomException(errors=self.logging)
    
    def write_csv(self, target_name, data):
        
        logging.info(f'Write mode: "{self.write_mode}" in Target files: "{target_name}"')
        
        state = "failed"
        try:
            ## read csv file.
            start_row = 2
            with open(target_name, "r") as reader:
                csvin = csv.DictReader(reader, skipinitialspace=True)
                rows = {idx + start_row: value for idx, value in enumerate(csvin)}
            
                for idx in data:
                    _data = {columns: values for columns, values in data[idx].items() if columns != "remark"}
                    
                    if  str(idx) in self.change_rows.keys() and data[idx]["remark"] in ["Updated", "Inserted"]:
                        ## update / insert rows.
                        logging.info(f'"{data[idx]["remark"]}" Rows:"({idx})" in Target files. Changed: "{self.change_rows[str(idx)]}"')
                        rows.update({idx: _data})
                    else:
                        if idx in self.remove_rows:
                            continue
                        ## no change rows.
                        rows[idx].update(_data)
                        
            ## write csv file.
            with open(target_name, 'w', newline='') as writer:
                csvout = csv.DictWriter(writer, csvin.fieldnames, delimiter=',', quotechar='"')
                csvout.writeheader()
                for idx in rows:
                    ## set data types column.
                    rows[idx]["CreateDate"] = rows[idx]["CreateDate"].strftime("%Y%m%d%H%M%S")
                    rows[idx]["LastUpdatedDate"] = rows[idx]["LastUpdatedDate"].strftime("%Y%m%d%H%M%S")
                    
                    if idx not in self.remove_rows:
                        csvout.writerow(rows[idx])
                        
            writer.closed 
            state = "succeed"
            
        except Exception as err:
            raise Exception(err)
        
        return state
    
    def validation_data(self, df: pd.DataFrame, change_df: pd.DataFrame) -> dict:

        logging.info("Verify Changed information..")
        self.logging[-1].update({"function": "validation_data"})

        self.change_rows = {}
        self.remove_rows = []
        if len(df.index) > len(change_df.index):
            self.remove_rows = [idx for idx in list(df.index) if idx not in list(change_df.index)]

        ## reset index data.
        union_index = np.union1d(df.index, change_df.index)
        ## target / tmp data.
        df = df.reindex(index=union_index, columns=df.columns).iloc[:, :-1]
        ## change data.
        change_df = change_df.reindex(index=union_index, columns=change_df.columns).iloc[:, :-1]
        
        df["count_change"] = pd.DataFrame(np.where(df.ne(change_df), True, False), index=df.index, columns=df.columns)\
            .apply(lambda x: (x == True).sum(), axis=1)

        def format_record(record):
            return ("{"+"\n".join("{!r}: {!r},".format(columns, values) \
                for columns, values in record.items())+"}")

        start_rows = 2
        for idx in union_index:
            if idx not in self.remove_rows:

                record = {}
                for data, change_data in zip(df.items(), change_df.items()):
                    if df.loc[idx, "count_change"] != 14:
                        ## No_changed rows.
                        if df.loc[idx, "count_change"] < 1:  # <=1
                            df.at[idx, data[0]] = data[1].iloc[idx]
                            df.loc[idx, "remark"] = "No_changed"

                        else:
                            ## Updated rows.
                            if data[1][idx] != change_data[1][idx]:
                                record.update({data[0]: f"{data[1][idx]} -> {change_data[1][idx]}"})
                            df.at[idx, data[0]] = change_data[1].iloc[idx]
                            df.loc[idx, "remark"] = "Updated"

                    else:
                        ## Inserted rows.
                        record.update({data[0]: change_data[1][idx]})
                        df.at[idx, data[0]] = change_data[1].iloc[idx]
                        df.loc[idx, "remark"] = "Inserted"

                if record != {}:
                    self.change_rows[start_rows + idx] = format_record(record)

            else:
                ## Removed rows.
                df.loc[idx, "remark"] = "Removed"
        self.remove_rows = [idx + start_rows for idx in self.remove_rows]

        df = df.drop(["count_change"], axis=1)
        df.index += start_rows
        _data = df.to_dict(orient='index')
    
        self.logging[-1].update({"status": "verify"})
        return _data
    
    def customize_data(self, target_df: pd.DataFrame, change_df: pd.DataFrame) -> dict:

        logging.info("Customize Data to Target..")
        self.logging[-1].update({"function": "customize_data"})
        
        date = self.fmt_batch_date
        data = {}
        try:
            ## filter data on batch date
            date_df = target_df[target_df["CreateDate"].isin(np.array([date])\
                .astype("datetime64[ns]"))].reset_index(drop=True)
            
            _data = self.validation_data(date_df, change_df)
            
            ## filter data not on batch date
            df = target_df[~target_df["CreateDate"].isin(np.array([pd.Timestamp(date)])\
                .astype("datetime64[ns]"))].iloc[:, :-1].to_dict("index")
            
            ## merge new data / old data
            max_rows = max(df, default=0)
            for idx, values in _data.items():
                if idx in self.change_rows or idx in self.remove_rows:
                    values.update({"mark_rows": idx})
                df = {**df, **{max_rows + idx: values}}
                
            ## sorted batch date order.
            start_rows = 2
            for idx, values in enumerate(sorted(df.values(), key=lambda d: d["CreateDate"])):
                idx += start_rows
                i = 0
                if "mark_rows" in values.keys():
                    if values["mark_rows"] in self.change_rows:
                        self.change_rows[str(idx)] = self.change_rows.pop(values["mark_rows"])
                    elif values["mark_rows"] in self.remove_rows:
                        self.remove_rows[i] = idx
                        i += 1
                    values.pop("mark_rows")
                data.update({idx: values})
            
        except Exception as err:
            raise Exception(err)
        
        return data

